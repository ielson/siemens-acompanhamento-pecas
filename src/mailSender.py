# coding=utf8


import pandas as pd

from email.message import EmailMessage
from email.headerregistry import Address
from email.utils import make_msgid
import datetime
import smtplib
import ssl

import dateutil.parser as dparser

import os


# vou usar o soup.find o soup.new_tag e o soup.append
from bs4 import BeautifulSoup

import html

import argparse
import logging
import logging.handlers
import csv
import collections
import statistics


from openpyxl import Workbook
from openpyxl import load_workbook

import matplotlib.pyplot as plt


from pandas.plotting import register_matplotlib_converters
from PyQt5 import QtWidgets

import predefinicoes

class QTextEditLogger(logging.Handler):
    def __init__(self, parent):
        super().__init__()
        self.widget = QtWidgets.QPlainTextEdit(parent)
        self.widget.setReadOnly(True)

    def emit(self, record):
        msg = self.format(record)
        self.widget.appendPlainText(msg)

def argParsing():
    parser = argparse.ArgumentParser()
    parser.add_argument('-l', '--loglevel', help="define the log level", default="info")
    args = parser.parse_args()
    return args


def get_loggers(parserLevel):
    formatterFull = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
    formatterHtml = logging.Formatter('%(message)s')

    if not os.path.exists('..\\Logs'):
        os.makedirs('..\\Logs')
    if not os.path.exists('..\\Logs\\logfile.log'):
        with open('..\\Logs\\logfile.log', 'w'): pass

    basic_handler = logging.handlers.RotatingFileHandler('..\Logs\logfile.log', maxBytes=512 * 1024, backupCount=1)
    basic_handler.setFormatter(formatterFull)
    basic_logger = logging.getLogger('basic_logger')
    basic_logger.setLevel(parserLevel)
    basic_logger.addHandler(basic_handler)


    # define a Handler which writes INFO messages or higher to the sys.stderr
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    formatterConsole = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    console.setFormatter(formatterConsole)
    logging.getLogger('basic_logger').addHandler(console)

    html_handler = logging.FileHandler('..\Logs\html_logfile.html', mode='w')
    html_handler.setFormatter(formatterHtml)
    html_logger = logging.getLogger('html_logger')
    html_logger.setLevel(parserLevel)
    html_logger.addHandler(html_handler)



    return basic_logger, html_logger


def geradorHistorico(planilha, pecas_por_eng):
    basic_logger.info("Atualizando arquivo de histórico")
    QtWidgets.QApplication.processEvents()
    # alterar esse jeito de inserir a media pra um melhor
    num_pecas_por_eng = pecas_por_eng.copy()
    item1 = num_pecas_por_eng.popitem()
    pecas_por_eng["Média"] = statistics.mean(list(num_pecas_por_eng.values()))
    pecas_por_eng["Total"] = sum(list(num_pecas_por_eng.values()))

    ultima_linha = planilha.max_row
    engs_registrados = {}
    for coluna in planilha.iter_cols(min_row=1, max_row=1, max_col=planilha.max_column, min_col=1):
        celula = coluna[0]
        engs_registrados[celula.value] = celula.column
    for next_col in range(1, len(CSES)+1):
        eng, quant = pecas_por_eng.popitem()
        if eng in engs_registrados:
            planilha.cell(column=engs_registrados[eng], row=ultima_linha+1, value=quant)
        else:
            # se o arquivo estiver vazio, corrigimos o ponto inicial
            # preciso achar um jeito melhor de ver se está vazio
            if planilha['A1'].value is None:
                ultima_coluna = 0
                basic_logger.debug("O histórico estava vazio")
            else:
                ultima_coluna = planilha.max_column
                basic_logger.debug("O histórico não estava vazio")

            # se nao tiver nada no arquivo ainda, comecamos da primeira linha
            planilha.cell(column=ultima_coluna+1, row=1, value=eng)
            planilha.cell(column=ultima_coluna+1, row=ultima_linha+1, value=quant)
    basic_logger.info("Arquivo de histórico atualizado")
    QtWidgets.QApplication.processEvents()

def plotter(caminhoHistorico):
    # precisa estar aqui para converter o datetime para ser plotado no matplotlib
    register_matplotlib_converters()
    # o que usei no interativo
    plt.style.use('seaborn-dark')
    fig, ax = plt.subplots(figsize=(8, 4.5))
    historico = pd.read_excel(caminhoHistorico)
    historico.fillna(0, inplace=True)
    objetivo = [1 for i in range(len(historico["Data"]))]

    for CSE in historico:
        if (CSE != "Data") and (CSE != "Média") and (CSE != "Total"):
            basic_logger.info("Gerando gráfico para {}".format(CSE))
            QtWidgets.QApplication.processEvents()
            ax.set_title('Evolução do CSE com relação ao num de peças')
            ax.set_xlabel('Data')
            ax.set_ylabel('Num de pecas')
            plotMedia = ax.plot(historico["Data"], historico["Média"], 'b-', linewidth=3, label='Média da Regional')
            plotObj = ax.plot(historico["Data"], objetivo, 'g--', linewidth=3, label='Objetivo de Pecas por Eng')
            plotCSE = ax.plot(historico["Data"], historico[CSE], 'r', linewidth=3, label='Peças atrasadas do Eng')
            ax.legend()

            #plt.text(historico["Data"].iloc[-1], historico[CSE].iloc[-1], 'Peças CSE', color=plotCSE[0].get_color())

            # ax.annotate("Peças CSE", xy=(historico["Data"].iloc[-1], historico["Dalmo"].iloc[-1]))
            #plt.text(historico["Data"].iloc[-1], historico["Média"].iloc[-1], 'Média Regional',color=plotMedia[0].get_color())
            #plt.text(historico["Data"].iloc[-1], objetivo[-1], 'Objetivo', color=plotObj[0].get_color())

            plt.ylim(0)
            fig.savefig(os.path.join(pastaGraficos, '{}.png'.format(CSE)), bbox_inches='tight')
            plt.cla()

    # geração de gráficos do gestor
    basic_logger.info("Gerando gráficos do gestor")
    QtWidgets.QApplication.processEvents()
    plotRegional = ax.plot(historico["Data"], historico["Total"])
    ax.set_title('Evolução do num de peças da Regional')
    ax.set_xlabel('Data')
    ax.set_ylabel('Num Peças')
    # preparando para a regressão linear
    """
    basic_logger.info("Gerando regressão linear")
    z = np.polyfit(historico["Data"], historico["Total"], 1)
    p = np.poly1d(z)
    plt.plot(historico["Data"], p(historico["Data"], 'r--'))
    """
    fig.savefig(os.path.join(pastaGraficos, 'gestor1.png'), bbox_inches='tight')
    plt.cla()

    cam = (213343.58, 171195.11)
    sup = (2097092.61, 1991775.23)
    sul = (370195.61, 570762.64)
    rio = (231694.96, 461844.69)
    cen = (1594293.54, 1265687.46)
    nor = (687483.88, 927045.98)
    ccc = (1228586.20, 1606814.25)

    varRegionais = {
        'CAM':(100*cam[1]-cam[0])/cam[0],
        'SUP':(100*sup[1]-sup[0])/sup[0],
        'SUL':(100*sul[1]-sul[0])/sul[0],
        'RIO':(100*rio[1]-rio[0])/rio[0],
        'CEN':(100*cen[1]-cen[0])/cen[0],
        'NOR':(100*nor[1]-nor[0])/nor[0],
        'CCC':(100*ccc[1]-ccc[0])/ccc[0]
    }


    plt.bar(range(len(varRegionais)), list(varRegionais.values()), align='center')
    plt.title('Variação percentual das regionais em relação ao custo de peças pendentes')
    plt.xlabel('Regional')
    plt.ylabel('Variação percentual')
    plt.xticks(range(len(varRegionais)), list(varRegionais.keys()))
    fig.savefig(os.path.join(pastaGraficos, 'gestor2.png'), bbox_inches='tight')
    plt.cla()

    plotRegionais = ax.plot("")

    return historico




class Colaborador:
    '''Classe que guarda os parametros de cada eng'''
    def __init__(self, codigo, nome, pecas):
        self.codigo = codigo
        self.nome = nome
        #self.email = predefinicoes.emails_dict[codigo]
        self.email = listaEmails.loc[codigo]
        self.pecas = pecas
        self.numPecas = len(pecas)

    def get_history(self):
        pass

    def get_graph(self):
        return os.path.join(pastaGraficos, '{}.png'.format(self.nome))

def mail_sender(colaborador, server):
    message = EmailMessage()
    message["Subject"] = "Acompanhamento de Peças"
    message["From"] = Address("Daniel do acompanhamento de peças", sender_email, "gmail.com")
    message["To"] = colaborador.email
    tabelaDePecas = colaborador.pecas[["Batch ID", "Ordem", "Texto breve material", "Ações Necessárias", "Dias em Campo", "Nome 1"]].to_html(
        index=False, justify="center")

    soup = BeautifulSoup(tabelaDePecas, 'html.parser')
    linhasPecas = soup.find('tbody').find_all('tr')
    for linha in linhasPecas:
        solicitacaoColeta = predefinicoes.html_pedir_coleta.format(batchId=linha.td.string, cliente=linha.find_all("td")[-1].string, chamado=linha.find_all("td")[1].string)
        html_logger.debug("solicitação: {}".format(solicitacaoColeta))
        linha.append(solicitacaoColeta)
        conteudo = predefinicoes.html_reportar_problema.format(numeroPeca=(linha.td.string))
        html_logger.debug("conteudo: {}".format(conteudo))
        linha.append(conteudo)

    tabelaDePecas = html.unescape(soup.prettify())
    html_logger.debug("tabela: {}".format(tabelaDePecas))

    custo = "235.840,57"

    image_cid = make_msgid(domain='xyz.com')
    message.set_content(predefinicoes.texto_email.format(name=colaborador.nome, num=colaborador.numPecas, pecas=colaborador.pecas[
        ["Material", "Texto breve material", "Ações Necessárias", "Dias em Campo", "Nome 1"]].to_string()))
    # codigoHtmlFormatado = codigoHtml.format(name=colaborador.nome, num=colaborador.numPecas, pecas=tabelaDePecas,
    #                                         custo=custo, image_cid=image_cid[1:-1])
    codigoHtmlFormatado = predefinicoes.html_email.format(name=colaborador.nome, num=colaborador.numPecas, pecas=tabelaDePecas,
                                            custo=custo, image_cid=image_cid[1:-1])

    message.add_alternative(codigoHtmlFormatado, subtype="html")

    with open(colaborador.get_graph(), 'rb') as arquivo:
        imagem = arquivo.read()
        message.add_attachment(imagem, maintype='image', subtype='png', filename="historico.png", cid=image_cid)

    html_logger.debug(codigoHtmlFormatado)

    server.send_message(message)
    basic_logger.info("Email enviado")
    basic_logger.info("Para: {}, {}, encontradas {} peças".format(colaborador.nome, colaborador.email, colaborador.numPecas))
    QtWidgets.QApplication.processEvents()
    del message


def mail_gestor(colaborador, historico, server):
    basic_logger.info("Criando email do gestor")
    message = EmailMessage()
    message["Subject"] = "Acompanhamento de Peças"
    message["From"] = Address("Daniel do acompanhamento de peças", sender_email, "gmail.com")
    message["To"] = colaborador.email

    maioresDevedores = sorted(dictCSES.values(), key=lambda item: item.numPecas, reverse=True)[:3]
    quantDevedores = [str(pessoa.numPecas) for pessoa in maioresDevedores]
    maioresDevedores = [pessoa.nome for pessoa in maioresDevedores]
    s = ", "
    maioresDevedores = s.join(maioresDevedores)
    quantDevedores = s.join(quantDevedores)
    basic_logger.info("Encontrando maiores devedores")
    QtWidgets.QApplication.processEvents()

    nenhumaPeca = historico.iloc[-1][historico.iloc[-1] == 0]
    nenhumaPeca = nenhumaPeca.index.tolist()
    nenhumaPeca = s.join(nenhumaPeca)
    basic_logger.info("Encontrando os que não devem nada")
    QtWidgets.QApplication.processEvents()

    deltaPercento = 100 * (historico["Total"].iloc[-1] - historico["Total"].iloc[-2]) / historico["Total"].iloc[-2]


    primeiro_cid = make_msgid(domain='xyz.com')
    segundo_cid = make_msgid(domain='xyz.com')
    message.set_content(
        predefinicoes.texto_gestor.format(name=colaborador.nome, variacao=deltaPercento, total=historico['Total'].iloc[-1],
                                            devedores=maioresDevedores, naodeve=nenhumaPeca, quantDevedores=quantDevedores))
    codigoHtmlFormatado = predefinicoes.html_gestor.format(name=colaborador.nome, variacao=deltaPercento,
            total=historico['Total'].iloc[-1], devedores=maioresDevedores, naodeve=nenhumaPeca, image_cid=primeiro_cid[1:-1], quantDevedores=quantDevedores, image_cid2=segundo_cid[1:-1])
    message.add_alternative(codigoHtmlFormatado, subtype="html")

    with open('..\\Gráficos\\gestor1.png', 'rb') as arquivo:
        imagem = arquivo.read()
        message.add_attachment(imagem, maintype='image', subtype='png', filename="historico.png", cid=primeiro_cid)

    with open('..\\Gráficos\\gestor2.png', 'rb') as arquivo:
        imagem = arquivo.read()
        message.add_attachment(imagem, maintype='image', subtype='png', filename="historico.png", cid=segundo_cid)

    html_logger.debug(codigoHtmlFormatado)

    #server.send_message(message)
    basic_logger.info("Email do gestor enviado")
    basic_logger.info(
        "Para: {}, {}".format(colaborador.nome, colaborador.email))
    QtWidgets.QApplication.processEvents()
    del message


def main(planilha):
    try:
        global CSES

        basic_logger.info("Inicializando o programa")
        print('dir padra: {}'.format(diretorioPadrao))
        print('pasta graficos: {}'.format(pastaGraficos))
        print('caminho emails: {}'.format(caminhoEmails))


        QtWidgets.QApplication.processEvents()
        if not os.path.exists(pastaGraficos):
            os.mkdir(pastaGraficos)
            print('criando pasta')
        else:
            print('pasta existe')
        print('pos pasta')
        codigoHtml = predefinicoes.html_email

        reportLogistica = pd.read_excel(planilha, header=0)
        reportLogisticaNor = reportLogistica[reportLogistica.Regional == 'MED-NOR']
        reportLogisticaNor10Mais = reportLogisticaNor[reportLogisticaNor["Dias em Campo"] > 10]
        pecasPorCSE = reportLogisticaNor10Mais.groupby('CSE Responsável')

        QtWidgets.QApplication.processEvents()
        filename = planilha.split('/')[-1]
        date = dparser.parse(filename, fuzzy=True, dayfirst=True, ignoretz=True)
        # convertendo datetime em só date
        date = date.date()
        #date = "{}/{}".format(planilha.split()[4].split('_')[0], planilha.split()[4].split('_')[1])
        CSES = collections.OrderedDict()
        CSES["Data"] = date
        basic_logger.info("Lendo planilha gerada em {}".format(date))
        QtWidgets.QApplication.processEvents()

        # abrir lista de emails
        global listaEmails
        listaEmails = pd.read_excel(caminhoEmails, index_col=0)
        global dictCSES
        dictCSES = {}

        # gera objeto dos CSES
        for CSE in pecasPorCSE.groups:
            nome = CSE.split("- ")[1].split()[0].capitalize()
            dictCSES[nome] = Colaborador(CSE, nome, pecasPorCSE.get_group(CSE))
            # esse CSES não precisa existir
            CSES[nome] = dictCSES[nome].numPecas
        gestor = Colaborador("Eduardo Zulli", "Zulli", "")

        # Preciso dar o reverse para quando usar o popitem ele ficar na ordem correta
        CSES = collections.OrderedDict(reversed(list(CSES.items())))

        historicoPath = "..\Histórico\Histórico.xlsx"
        if not os.path.exists('..\\Histórico'):
            os.makedirs('..\\Histórico')

        if not os.path.exists(historicoPath):
            # cria o arquivo se ele não existir
            wb = Workbook()
            wb.save(historicoPath)

        wb = load_workbook(historicoPath)
        planilha = wb.active
        datas = []
        for linha in planilha.iter_rows(min_row=1, max_col=1):
            celula = linha[0]
            if type(celula.value) == datetime.datetime:
                datas.append(celula.value.date())
        if CSES['Data'] not in datas:
            # TODO Colocar data sempre em primeiro na planilha e não procurar pelo primeiro, mas sim por data
            # TODO Tratar erro de quando o historico tá aberto
            basic_logger.info("Gerando histórico dos CSES")
            QtWidgets.QApplication.processEvents()
            geradorHistorico(planilha, CSES)
        else:
            basic_logger.info("Data já consta no histórico. Não gerarei nova entrada")
            QtWidgets.QApplication.processEvents()
        wb.save(historicoPath)

        historic = plotter(historicoPath)

        context = ssl.create_default_context()
        s = smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context)
        s.set_debuglevel(False)
        with s as server:
            basic_logger.info("Fazendo login no servidor de email")
            server.login(sender_email, password)
            for colaborador in dictCSES.values():
                mail_sender(colaborador, server)
            mail_gestor(gestor, historic, server)
    except AttributeError:
        basic_logger.exception("Formato de arquivo não corresponde ao report de peças esperado.")
        basic_logger.warning("Favor escolher outro arquivo.")
    except KeyError as cseNaoEncontrado:
        basic_logger.error("O CSE {} não foi encontrado".format(cseNaoEncontrado))
        basic_logger.warning("Favor verificar o nome na lista de emails")
    except Exception as error:
        print(error)
sender_email = "parts.nord"
password = "partsnord2020"
diretorioPadrao =  os.path.abspath('..')
print('dir padra: {}'.format(diretorioPadrao))
pastaGraficos = os.path.join(diretorioPadrao, 'graficos')
print('pasta graficos: {}'.format(pastaGraficos))
caminhoEmails = os.path.join(diretorioPadrao, 'lista_de_emails.xlsx')
print('caminho emails: {}'.format(caminhoEmails))

if __name__ == '__main__':
    planilha = '../Histórico/'
    planilha += 'Report Peças em campo 30_10 - SAS.xlsx'
    args = argParsing()
    parserLevel = getattr(logging, args.loglevel.upper(), None)
    if not isinstance(parserLevel, int):
        raise ValueError('Invalid log level: {}'.format(args.loglevel))
    basic_logger, html_logger = get_loggers(parserLevel)
    main(planilha)
else:
    parserLevel = 1
    basic_logger, html_logger = get_loggers(parserLevel)



